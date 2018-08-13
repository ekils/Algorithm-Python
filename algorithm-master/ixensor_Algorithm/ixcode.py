#coding=utf-8

import os
import threading
from playsound import playsound
from PIL import Image , ImageDraw
import  numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import pandas as pd
import re
from colour import Color
import statistics
from decimal import Decimal
import linecache  # 讀取特定行數，但是檔案必須要給絕對路徑
from openpyxl import load_workbook    # 這邊引入是要能夠讀取excel的row數量


class ixensor:

    def __init__(self):
        self.CodeDone = ('/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/smb_stage_clear.mp3')

    def sound(self):
        playsound(self.CodeDone)

    # 均勻度分析
    def Uniformity(self):

        ori_path= os.getcwd()
        choice= input("Choose (1.x588 2.iphone ) (1.NOTM 2.YESTM) : ")
        folder= choice[0]
        tone_mapping= choice[1]

        if folder=='1':
            if tone_mapping =='1':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/x588/{}".format('NOTM'))
            elif tone_mapping=='2':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/x588/{}".format('YESTM'))

        elif folder =="2":
            if tone_mapping == '1':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/iphone/{}".format('NOTM'))
            elif tone_mapping == '2':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/iphone/{}".format('YESTM'))


        img= Image.open("Initial_01.png")
        img_array = np.asarray(img)
        r= img_array[50:650,100:1000,0]

        print(r.shape)

        x= np.arange(0,900)
        y= np.arange(0,600)
        X, Y = np.meshgrid(x, y)
        z= r
        levels = np.arange(100,255,15)
        CS = plt.contour(X, Y, z, levels=levels)
        plt.clabel(CS, inline=3, fontsize=13)
        plt.colorbar()
        add_thread = threading.Thread(target=self.sound)   # 增加線程
        add_thread.start()
        plt.show()
        img.close()

        os.chdir(ori_path)


    def DarkSignal(self):
        folder = input('before or after: ')
        origin = os.getcwd()
        os.chdir('/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/{}'.format(folder))

        r_10_index = []
        if folder == 'before':
            for num in range(15, 45):
                # 先讀圖
                before = "DarkSignal_00{}.png".format(num)
                img = Image.open(before)
                print(img.mode, img.size)

                # 標註點
                r_add = 0
                draw = ImageDraw.Draw(img)
                for i in range(40):
                    coords = [(550 + i, 340)]
                    for j in range(40):
                        coords = [(550 + i, 340 + j)]
                        for (x, y) in coords:
                            r, g, b = img.getpixel((x, y))
                            # print(r)
                            r_add = r_add + r

                r_avg = float((r_add) / (40 * 40))
                print(r_avg)
                r_10_index.append(r_avg)


        elif folder == 'after':
            for num in range(785, 815):
                # 先讀圖
                before = "DarkSignal_01{}.png".format(num)
                img = Image.open(before)
                print(img.mode, img.size)

                # 標註點
                r_add = 0
                draw = ImageDraw.Draw(img)
                for i in range(40):
                    coords = [(550 + i, 340)]
                    for j in range(40):
                        coords = [(550 + i, 340 + j)]
                        for (x, y) in coords:
                            r, g, b = img.getpixel((x, y))
                            r_add = r_add + r

                r_avg = float((r_add) / (40 * 40))
                print(r_avg)
                r_10_index.append(r_avg)

        # 標準差
        frame_10_avg = round(np.mean(r_10_index), 4)
        frame_10_sd = round(np.std(r_10_index), 4)
        # frame_10_cv=  (frame_10_sd/frame_10_avg)*100
        print("frame_10_avg: {}".format(frame_10_avg))
        print("frame_10_sd: {}".format(frame_10_sd))
        # print ("frame_10_cv: {}%".format(frame_10_cv))


        # 作圖
        X = range(len(r_10_index))
        Y = r_10_index
        plt.title('Dark Signal(Area 40*40)')
        plt.xlabel('frame')
        plt.ylabel('R_signal avg ')
        plt.plot(X, Y)
        add_thread = threading.Thread(target=self.sound)    # 增加線程
        add_thread.start()
        plt.show()
        #img.close()

        os.chdir(origin)

    # 均勻度分析 3D
    def Contour_3d(self):


        ori_path= os.getcwd()
        choice= input("Choose (1.x588 2.iphone ) (1.NOTM 2.YESTM) : ")
        folder= choice[0]
        tone_mapping= choice[1]

        if folder=='1':
            if tone_mapping =='1':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/x588/{}".format('NOTM'))
            elif tone_mapping=='2':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/x588/{}".format('YESTM'))

        elif folder =="2":
            if tone_mapping == '1':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/iphone/{}".format('NOTM'))
            elif tone_mapping == '2':
                os.chdir("/Users/bobobo746/Desktop/stuff/python/Python3/ixensor_Algorithm/iphone/{}".format('YESTM'))


        img= Image.open("Initial_01.png")
        img_array = np.asarray(img)
        r= img_array[50:650,100:1000,0]

        print(r.shape)

        x= np.arange(0,900)
        y= np.arange(0,600)
        X, Y = np.meshgrid(x, y)
        z= r
        levels = np.arange(100,255,15)

        # 畫3d :
        fig= plt.figure()
        ax= Axes3D(fig)
        ax.plot_surface(X,Y,z,rstride=40,cstride=10,cmap=plt.get_cmap('inferno'))
        #ax.contourf(X,Y,z,zdir='z',cmap='inferno')



        add_thread = threading.Thread(target=self.sound)   # 增加線程
        add_thread.start()
        plt.show()
        img.close()

        os.chdir(ori_path)


    def Folder_Sort(self):
        ori_path = os.getcwd()
        ori_path=str(ori_path)
        file = 'Image_0001.png'
        # 讀取excel ,依照sheet
        named_r = pd.read_excel('X588.xlsx', sheetname='for_r')
        named_g = pd.read_excel('X588.xlsx', sheetname='for_g')
        named_b = pd.read_excel('X588.xlsx', sheetname='for_b')

        rgb_folder= ['r','g','b']
        for i in rgb_folder:
            path1= ori_path + '/' +i
            os.chdir(path1)  # 依序進入 r,g.b 資料夾
            if i=='r':
                for path, folder, allname in os.walk(path1):   # os.walk 獲取順序為：路徑,資料夾,檔案
                    if folder:  # 如果不是空檔案
                        for index, names in enumerate(folder):
                            os.rename(names,str(named_r['r'][index])) # 修改檔名


                    for path, folder, allname in os.walk(path1):

                        for ii in folder: # 修改後folder的檔名重新獲取
                            for path, folders, allname in os.walk(path1+'/'+ii):
                                for tt in allname:
                                    print(tt)
                                    if tt != file:
                                        if tt == '.DS_Store':
                                            print ('*')
                                        else:
                                            print('delete')
                                            os.remove(path + '/' + tt)
                                    else:
                                        print('keep')
                        os.chdir(path1)

            elif i=='g':
                for path, folder, allname in os.walk(path1):
                    if folder:
                        for index, names in enumerate(folder):
                            os.rename(names,str(named_g['g'][index]))


                    for path, folder, allname in os.walk(path1):

                        for ii in folder: # 修改後folder的檔名重新獲取
                            for path, folders, allname in os.walk(path1+'/'+ii):
                                for tt in allname:
                                    print(tt)
                                    if tt != file:
                                        if tt == '.DS_Store':
                                            print ('*')
                                        else:
                                            print('delete')
                                            os.remove(path + '/' + tt)
                                    else:
                                        print('keep')
                        os.chdir(path1)


            elif i=='b':
                for path, folder, allname in os.walk(path1):
                    if folder:
                        for index, names in enumerate(folder):
                            os.rename(names,str(named_b['b'][index]))


                    for path, folder, allname in os.walk(path1):

                        for ii in folder: # 修改後folder的檔名重新獲取
                            for path, folders, allname in os.walk(path1+'/'+ii):
                                for tt in allname:
                                    print(tt)
                                    if tt != file:
                                        if tt == '.DS_Store':
                                            print ('*')
                                        else:
                                            print('delete')
                                            os.remove(path + '/' + tt)
                                    else:
                                        print('keep')
                        os.chdir(path1)

            else:
                print('Something Wrong')
            os.chdir(ori_path)

        # 過曝比較
        def over_exposure_comapre(self):
            ori_path = os.getcwd()
            os.chdir(ori_path + '/' + 'NG')
            path2 = os.getcwd()
            for l in os.listdir(path2):
                if l != ".DS_Store":
                    print('Folder: {}'.format(l))
                    for ll in os.listdir(path2 + '/' + l):
                        if ll != ".DS_Store":
                            print(ll)
                            os.chdir(os.getcwd() + '/' + str(l) + '/' + str(ll))
                            block_list = []

                            if os.path.exists(os.getcwd() + '/' + 'Blue_Light_001.jpg'):
                                img = Image.open('Blue_Light_001.jpg').convert('L')
                                print(img.mode, img.size)

                                rec_x_ist = []
                                rec_y_ist = []

                                # 計算平均值
                                rec_x0 = 349
                                rec_y0 = 272
                                for ii in [136]:  # y
                                    for i in [100, 100 * 3.5]:  # x
                                        rec_x_ist.append(rec_x0 + i)
                                        rec_y_ist.append(rec_y0 + ii)

                                for kk in range(len(rec_x_ist)):
                                    block_add = 0
                                    for i in range(40 * 2):
                                        for j in range(40):
                                            coords = [(rec_x_ist[kk] + i, rec_y_ist[kk] + j)]
                                            for (x, y) in coords:
                                                block = img.getpixel((x, y))
                                                block_add = block_add + block

                                    block_avg = round(float((block_add) / (40 * 2 * 40)), 3)
                                    block_list.append(block_avg)
                                    print('[{}],{} '.format(kk + 1, block_avg))

                                if block_list[0] > block_list[1]:
                                    print('<<<<  light_leak , in which path: \n{}  >>>>>>'.format(
                                        os.getcwd() + '/' + l + '/' + ll))

                        os.chdir(path2)

            os.chdir(ori_path)

        # 訊號曲線圖
        def Signal_curve_A1C_lIPID(self):
            ori_path = os.getcwd()
            # 給定選定資料夾：
            folder = 'FOLDER'
            filename = 'DataRecord'
            # 給定張數：
            total_pic = 10
            file_num = [f for f in range(1, total_pic + 1)]

            # 圖框大小
            plt.figure(figsize=(12, 5))
            # Color 色塊範圍
            cc = Color("firebrick")
            colors = list(cc.range_to(Color("mediumpurple"), total_pic))  # 看幾張圖給幾個色
            for j, i in enumerate(file_num):
                with open(ori_path + '/' + folder + '/' + filename + str(i) + '.txt', 'r') as f:
                    ROI1_position = [index for index, label in enumerate(f.readlines()) if label == '$ROI 1 R\n']
                    f.close()
                # print(ROI1_position[0])
                with open(ori_path + '/' + folder + '/' + filename + str(i) + '.txt', 'r') as f:
                    p = [i.strip() for i in f.readlines()]
                    p[ROI1_position[0] + 1].split(',')
                    f.close()
                p = p[ROI1_position[0] + 1].split(',')
                y = [float(kk) for kk in p]
                x = [i for i in range(len(y))]
                # print(y)
                plt.plot(x, y, c='{}'.format(colors[j]), alpha=1.0, label='{}'.format(j))

            os.chdir(ori_path)
            plt.xlabel('Time')
            plt.ylabel('Signal')
            plt.title('Time_signal plot')
            plt.legend()
            # 畫格點
            plt.grid()
            plt.show()

            def Signal_Click(self):  # 算即時訊號

                ori_path = os.getcwd()
                os.chdir(ori_path + '/' + 'signal_click')

                # 讀圖 轉灰階
                img = Image.open("ProductionLineFiveColorCardProcessDone4.jpg").convert('L')
                print(img.mode, img.size)

                rec_x_ist = []
                rec_y_ist = []

                # 計算平均值
                rec_x0 = 349
                rec_y0 = 272
                for ii in [136]:  # y
                    for i in [100, 100 * 3.5]:  # x
                        rec_x_ist.append(rec_x0 + i)
                        rec_y_ist.append(rec_y0 + ii)

                for kk in range(len(rec_x_ist)):
                    block_add = 0
                    for i in range(40 * 2):
                        for j in range(40):
                            coords = [(rec_x_ist[kk] + i, rec_y_ist[kk] + j)]
                            for (x, y) in coords:
                                block = img.getpixel((x, y))
                                block_add = block_add + block

                    block_avg = round(float((block_add) / (40 * 2 * 40)), 3)
                    print('[{}],{} '.format(kk + 1, block_avg))
                imggray = img

                # 畫圖轉rgb:
                img2 = imggray.convert('RGB')

                print(img2.mode, img2.size)
                # 畫框
                draw = ImageDraw.Draw(img2)
                for ii in [136]:  # y
                    for i in [100, 100 * 3.5]:  # x
                        draw.rectangle([(rec_x0 + i, rec_y0 + ii), (rec_x0 + i + 40 * 2, rec_y0 + 40 + ii)],
                                       outline=(255, 0, 0))  # 紅框
                        rec_x_ist.append(rec_x0 + i)
                        rec_y_ist.append(rec_y0 + ii)

                # 秀圖
                img2.show()
                os.chdir(ori_path)



    # 五色卡：算八格訊號（連接白光分析白光畫圖）
    def FiveColorCard_Signal_Comfirm(self):

        # = == == =標準試片八區域排列順序 == == == =#
        #            5 - 6 - 7 - 8               #
        #            C - R - M - 4               #
        # = == == == == == == == == == == == == =#

        # = == == =分析試片八區域排列順序 == == == =#
        #            1 - 2 - 3 - 4               #
        #            5 - 6 - 7 - 8               #
        # = == == == == == == == == == == == == =#

        ori_path = os.getcwd()

        analysis_folder = ori_path + '/20180410'
        folder_path= os.listdir(analysis_folder)
        image_file = "ProductionLineFiveColorCardProcessDone.jpg"

        folder_count = 0

        zone_list = []

        all_zone_list = []
        for folder in folder_path:
            if folder!='.DS_Store':
                print('Now in folder: {}'.format(folder))
                os.chdir(analysis_folder + '/' +folder)

                # 讀圖
                if  os.path.exists(image_file):
                    img = Image.open(image_file)
                    # print(img.mode, img.size)

                    rec_x_ist = []
                    rec_y_ist = []

                    # 畫框
                    draw = ImageDraw.Draw(img)
                    rec_x0= 349
                    rec_y0= 272
                    for ii in [0, 136]: # y
                        for i in [0,133,133*2,133*3]:  # x
                            draw.rectangle([(rec_x0 + i,rec_y0 + ii), (rec_x0 + i + 40, rec_y0 + 40 + ii)], outline=(255, 100, 0))
                            rec_x_ist.append(rec_x0 + i)
                            rec_y_ist.append( rec_y0 + ii)

                    # 計算平均值
                    for kk in range(len(rec_x_ist)):
                        r_add = 0
                        g_add = 0
                        b_add = 0
                        for i in range(40):
                            for j in range(40):
                                coords = [(rec_x_ist[kk] + i, rec_y_ist[kk] + j)]
                                for (x, y) in coords:
                                    r, g, b = img.getpixel((x, y))
                                    # print(r)
                                    r_add = r_add + r
                                    g_add = g_add + g
                                    b_add = b_add + b

                        r_avg = round(float((r_add) / (40 * 40)),3)
                        g_avg = round(float((g_add) / (40 * 40)),3)
                        b_avg = round(float((b_add) / (40 * 40)),3)
                        print('[{}],{},{},{} '.format(kk+1,r_avg,g_avg,b_avg))
                        zone_list.append([r_avg,g_avg,b_avg])

                else:
                    print("Folder '{}' doesn't find image file. \n".format(folder))

                os.chdir(analysis_folder )
                # print(zone_list)
                all_zone_list.append(zone_list)
                zone_list = []
                folder_count += 1
        # 秀圖
        #img.show()

        self.FiveColorCard_WhiteLight_statictics( all_zone_list)
        os.chdir(ori_path)

    # 白光分析
    def FiveColorCard_WhiteLight_statictics(self,all_zone_list):
        print('\n')
        print('== == == 標準試片八區域排列順序 == == == \n'
              '           5 - 5 - 7 - 8               \n'
              '           C - R - M - 4                \n'
              '= == == == == == == == == == == == == =')


        print('== == == 分析試片八區域排列順序 == == == \n'
              '           1 - 2 - 3 - 4               \n'
              '           5 - 6 - 7 - 8               \n'
              '= == == == == == == == == == == == == =')

        list_r_avg = []
        list_g_avg = []
        list_b_avg = []
        plot_r_list= []
        plot_g_list= []
        plot_b_list= []
        plot_mean_r_list = []
        plot_mean_g_list = []
        plot_mean_b_list = []


        for ii in range(8): # zone 有八區
            for i in range(len(all_zone_list)):        # all_zone_list[0][0]  第一個是 folder, 第二個是 zone區
                list_r_avg.append(all_zone_list[i][ii][0])
                list_g_avg.append(all_zone_list[i][ii][1])
                list_b_avg.append(all_zone_list[i][ii][2])

            plot_r_list.append(list_r_avg)
            plot_g_list.append(list_g_avg)
            plot_b_list.append(list_b_avg)

            mean_r  = round(statistics.mean(list_r_avg),2)
            mean_g  = round(statistics.mean(list_g_avg),2)
            mean_b  = round(statistics.mean(list_b_avg),2)
            stdev_r = round(statistics.stdev(list_r_avg),2)
            stdev_g = round(statistics.stdev(list_g_avg),2)
            stdev_b = round(statistics.stdev(list_b_avg),2)
            cv_r = round(Decimal(stdev_r/mean_r*100),2)
            cv_g = round(Decimal(stdev_g/mean_g*100),2)
            cv_b = round(Decimal(stdev_b/mean_b*100),2)

            plot_mean_r_list.append(mean_r)
            plot_mean_g_list.append(mean_g)
            plot_mean_b_list.append(mean_b)

            list_r_avg = []
            list_g_avg = []
            list_b_avg = []

            print('第{}區：'.format(ii+1))
            print('[mean] {},{},{}'.format(mean_r,mean_g,mean_b))
            print('[stdev]{},{},{}'.format(stdev_r, stdev_g, stdev_b))
            print('[ cv%] {},{},{}'.format(cv_r,cv_g,cv_b))

        self.FiveColorCard_WhiteLight_plot( plot_b_list,plot_r_list,plot_mean_b_list,plot_mean_r_list)

    # 白光畫圖
    def FiveColorCard_WhiteLight_plot(self,plot_b_list,plot_r_list,plot_mean_b_list,plot_mean_r_list):
        # 畫圖：

        plt.figure(figsize=(12, 6))
        plt.figure(1)

        # zone 5:
        plt.subplot(241)  # 241 is right

        data = plot_b_list[0]
        bins = np.arange(60, 80, 0.5)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='b', alpha=0.7)
        plt.xlabel('Signal Value')
        plt.title('5')
        plt.grid(True)

        # zone 6:
        plt.subplot(242)

        data = plot_b_list[1]
        bins = np.arange(30, 50, 0.5)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='b', alpha=0.7)
        plt.xlabel('Signal Value')
        plt.title('6')
        plt.grid(True)

        # zone 7:
        plt.subplot(243)

        data = plot_r_list[2]
        bins = np.arange(130, 150, 0.5)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='r', alpha=0.7)
        plt.xlabel('Signal Value')
        plt.title('7')
        plt.grid(True)

        # zone 8
        plt.subplot(244)

        data = plot_r_list[3]
        bins = np.arange(70, 85, 1)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='r', alpha=0.7)
        plt.xlabel('Signal Value')
        plt.title('8')
        plt.grid(True)

        # zone C:
        plt.subplot(245)

        data = plot_r_list[4]
        bins = np.arange(115, 140, 0.5)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='k', alpha=0.5)
        plt.xlabel('Signal Value')
        plt.title('C')
        plt.grid(True)

        # zone R:
        plt.subplot(246)

        data = plot_r_list[5]
        bins = np.arange(150, 170, 0.5)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='k', alpha=0.5)
        plt.xlabel('Signal Value')
        plt.title('R')
        plt.grid(True)

        # zone M:
        plt.subplot(247)

        data = plot_r_list[6]
        bins = np.arange(170, 190, 0.5)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='k', alpha=0.5)
        plt.xlabel('Signal Value')
        plt.title('M')
        plt.grid(True)

        # zone 4:
        plt.subplot(248)

        data = plot_r_list[7]
        bins = np.arange(230, 250, 0.5)  # fixed bin size
        plt.xlim([min(data) - 1, max(data) + 1])
        plt.hist(data, bins=bins, color='k', alpha=0.5)
        plt.xlabel('Signal Value')
        plt.title('4')
        plt.grid(True)

        plt.subplots_adjust(hspace=0.5)
        # plt.show()



        # Scatter plot:
        plt.figure(figsize=(12, 6))
        plt.figure(2)
        # zone 5:
        plt.subplot(241)

        x = [i for i in range(len(plot_b_list[0]))]
        y = plot_b_list[0]
        plt.scatter(x, y, c='b', alpha=0.5)

        mean_y = [plot_mean_b_list[0]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('5')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        # zone 6:
        plt.subplot(242)

        x = [i for i in range(len(plot_b_list[1]))]
        y = plot_b_list[1]
        plt.scatter(x, y, c='b', alpha=0.5)

        mean_y = [plot_mean_b_list[1]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('6')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        # zone 7:
        plt.subplot(243)

        x = [i for i in range(len(plot_r_list[2]))]
        y = plot_r_list[2]
        plt.scatter(x, y, c='r', alpha=0.5)

        mean_y = [plot_mean_r_list[2]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('7')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        # zone 8:
        plt.subplot(244)

        x = [i for i in range(len(plot_r_list[3]))]
        y = plot_r_list[3]
        plt.scatter(x, y, c='r', alpha=0.5)

        mean_y = [plot_mean_r_list[3]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('7')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        # zone C:
        plt.subplot(245)

        x = [i for i in range(len(plot_r_list[4]))]
        y = plot_r_list[4]
        plt.scatter(x, y, c='k', alpha=0.5)

        mean_y = [plot_mean_r_list[4]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('C')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        # zone R:
        plt.subplot(246)

        x = [i for i in range(len(plot_r_list[5]))]
        y = plot_r_list[5]
        plt.scatter(x, y, c='k', alpha=0.5)

        mean_y = [plot_mean_r_list[5]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('R')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        # zone M:
        plt.subplot(247)

        x = [i for i in range(len(plot_r_list[6]))]
        y = plot_r_list[6]
        plt.scatter(x, y, c='k', alpha=0.5)

        mean_y = [plot_mean_r_list[6]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('M')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        # zone 4:
        plt.subplot(248)

        x = [i for i in range(len(plot_r_list[7]))]
        y = plot_r_list[7]
        plt.scatter(x, y, c='k', alpha=0.5)

        mean_y = [plot_mean_r_list[7]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)

        plt.title('4')
        plt.ylabel('Signal Value')
        plt.grid(True)
        plt.ylim([min(y) - 10, max(y) + 10])

        plt.subplots_adjust(hspace=0.5)
        plt.show()

    # 五色卡: 單色光分析
    def FiveColorCard_SingleLight_statictics(self):

        file_name = 'DataRecord.txt'

        ori_path = os.getcwd()
        analysis_folder = ori_path + '/20180511'
        folder_path= os.listdir(analysis_folder)

        Blue_Light_Cali = []
        Blue_Light_Reac = []
        Blue_Light_ROI5 = []
        Blue_Light_ROI6 = []

        Green_Light_Cali = []
        Green_Light_Reac = []
        Green_Light_ROI5 = []
        Green_Light_ROI6 = []

        Red_Light_Manu = []
        Red_Light_ROI4 = []
        Red_Light_ROI7 = []
        Red_Light_ROI8 = []

        for folder in folder_path:
            if folder!='.DS_Store':
                os.chdir(analysis_folder + '/' + folder)
                print('Now in folder: {}'.format(folder))

                if os.stat(file_name).st_size == 0:
                    print('====== 空空如也 ======')
                else:
                    count =0
                    # print(os.getcwd())
                    with open(file_name,'r') as f:
                        for i in f.readlines():
                            count += 1
                        # BLUE:
                            if '$Production_Line_Calibration_FiveColor_Blue_Light_Calibration X_Calibration Y' in i:
                                Blue_Light_Cali.append(linecache.getline(analysis_folder + '/' + folder +'/'+file_name, count+1)) # 需給定絕對路徑
                            if '$Production_Line_Calibration_FiveColor_Blue_Light_Reaction X_Reaction Y' in i:
                                Blue_Light_Reac.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))
                            if '$Production_Line_Calibration_FiveColor_Blue_Light_ROI5 X_ROI5 Y' in i :
                                Blue_Light_ROI5.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))
                            if '$Production_Line_Calibration_FiveColor_Blue_Light_ROI6 X_ROI6 Y' in i:
                                Blue_Light_ROI6.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))
                        # GREEN:
                            if '$Production_Line_Calibration_FiveColor_Green_Light_Calibration X_Calibration Y' in i:
                                Green_Light_Cali.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))  # 需給定絕對路徑
                            if '$Production_Line_Calibration_FiveColor_Green_Light_Reaction X_Reaction Y' in i:
                                Green_Light_Reac.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))
                            if '$Production_Line_Calibration_FiveColor_Green_Light_ROI5 X_ROI5 Y' in i:
                                Green_Light_ROI5.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))
                            if '$Production_Line_Calibration_FiveColor_Green_Light_ROI6 X_ROI6 Y' in i:
                                Green_Light_ROI6.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))

                        # RED:
                            if '$Production_Line_Calibration_FiveColor_Red_Light_Manual X_Manual Y' in i:
                                Red_Light_Manu.append(linecache.getline(analysis_folder + '/' + folder +'/'+file_name, count+1))
                            if '$Production_Line_Calibration_FiveColor_Red_Light_ROI4 X_ROI4 Y' in i:
                                Red_Light_ROI4.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))
                            if '$Production_Line_Calibration_FiveColor_Red_Light_ROI7 X_ROI7 Y' in i :
                                Red_Light_ROI7.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))
                            if '$Production_Line_Calibration_FiveColor_Red_Light_ROI8 X_ROI8 Y' in i:
                                Red_Light_ROI8.append(linecache.getline(analysis_folder + '/' + folder + '/' + file_name, count + 1))


            os.chdir(analysis_folder)
        Blue_Light_Cali = [i.strip() for i in Blue_Light_Cali]
        Blue_Light_Reac = [i.strip() for i in Blue_Light_Reac]
        Blue_Light_ROI5 = [i.strip() for i in Blue_Light_ROI5]
        Blue_Light_ROI6 = [i.strip() for i in Blue_Light_ROI6]

        Green_Light_Cali = [i.strip() for i in Green_Light_Cali]
        Green_Light_Reac = [i.strip() for i in Green_Light_Reac]
        Green_Light_ROI5 = [i.strip() for i in Green_Light_ROI5]
        Green_Light_ROI6 = [i.strip() for i in Green_Light_ROI6]

        Red_Light_Manu  = [i.strip() for i in Red_Light_Manu]
        Red_Light_ROI4  = [i.strip() for i in Red_Light_ROI4]
        Red_Light_ROI7  = [i.strip() for i in Red_Light_ROI7]
        Red_Light_ROI8  = [i.strip() for i in Red_Light_ROI8]

        zone5_C_Ratio_list_B = []
        zone6_R_Ratio_list_B = []
        zone5_C_Ratio_list_G = []
        zone6_R_Ratio_list_G = []
        zone7_M_Ratio_list_R = []
        zone8_4_Ratio_list_R = []


        for i in range(len(Blue_Light_Cali)):
            zone5_C_Ratio_list_B.append( round(float(Blue_Light_ROI5[i].split(',')[0]) / float(Blue_Light_Cali[i].split(',')[0]),3) )
            zone6_R_Ratio_list_B.append( round(float(Blue_Light_ROI6[i].split(',')[0]) / float(Blue_Light_Reac[i].split(',')[0]),3) )
            zone5_C_Ratio_list_G.append( round(float(Green_Light_ROI5[i].split(',')[1]) / float(Green_Light_Cali[i].split(',')[1]),3) )
            zone6_R_Ratio_list_G.append( round(float(Green_Light_ROI6[i].split(',')[1]) / float(Green_Light_Reac[i].split(',')[1]),3) )
            zone7_M_Ratio_list_R.append( round(float(Red_Light_ROI7[i].split(',')[2]) / float(Red_Light_Manu[i].split(',')[2]),3) )
            zone8_4_Ratio_list_R.append( round(float(Red_Light_ROI8[i].split(',')[2]) / float(Red_Light_ROI4[i].split(',')[2]),3) )

        # print('5_C_Ratio:\n{}'.format(zone5_C_Ratio_list))
        # print('6_R_Ratio:\n{}'.format(zone6_R_Ratio_list))
        # print('7_M_Ratio:\n{}'.format(zone7_M_Ratio_list))
        # print('8_4_Ratio:\n{}'.format(zone8_4_Ratio_list))

        mean_5_C_B = round(statistics.mean(zone5_C_Ratio_list_B), 3)
        mean_6_R_B = round(statistics.mean(zone6_R_Ratio_list_B), 3)
        mean_5_C_G = round(statistics.mean(zone5_C_Ratio_list_G), 3)
        mean_6_R_G = round(statistics.mean(zone6_R_Ratio_list_G), 3)
        mean_7_M_R = round(statistics.mean(zone7_M_Ratio_list_R), 3)
        mean_8_4_R = round(statistics.mean(zone8_4_Ratio_list_R), 3)

        stdev_5_C_B = round(statistics.stdev(zone5_C_Ratio_list_B), 3)
        stdev_6_R_B = round(statistics.stdev(zone6_R_Ratio_list_B), 3)
        stdev_5_C_G = round(statistics.stdev(zone5_C_Ratio_list_G), 3)
        stdev_6_R_G = round(statistics.stdev(zone6_R_Ratio_list_G), 3)
        stdev_7_M_R = round(statistics.stdev(zone7_M_Ratio_list_R), 3)
        stdev_8_4_R = round(statistics.stdev(zone8_4_Ratio_list_R), 3)

        cv_5_C_B = round(Decimal(stdev_5_C_B / mean_5_C_B * 100), 3)
        cv_6_R_B = round(Decimal(stdev_6_R_B / mean_6_R_B * 100), 3)
        cv_5_C_G = round(Decimal(stdev_5_C_G / mean_5_C_G * 100), 3)
        cv_6_R_G = round(Decimal(stdev_6_R_G / mean_6_R_G * 100), 3)
        cv_7_M_R = round(Decimal(stdev_7_M_R / mean_7_M_R * 100), 3)
        cv_8_4_R = round(Decimal(stdev_8_4_R / mean_8_4_R * 100), 3)


        print('==== 5_C_Ratio_B:  [Mean]   [Stdev]   [ CV% ] ====\n'
              '                    {}      {}     {}'.format(mean_5_C_B,stdev_5_C_B,cv_5_C_B))
        print('==== 6_R_Ratio_B:  [Mean]   [Stdev]   [ CV% ] ====\n'
              '                    {}      {}     {}'.format(mean_6_R_B,stdev_6_R_B,cv_6_R_B))

        print('==== 5_C_Ratio_G:  [Mean]   [Stdev]   [ CV% ] ====\n'
              '                    {}      {}     {}'.format(mean_5_C_G,stdev_5_C_G,cv_5_C_G))
        print('==== 6_R_Ratio_G:  [Mean]   [Stdev]   [ CV% ] ====\n'
              '                    {}      {}     {}'.format(mean_6_R_G,stdev_6_R_G,cv_6_R_G))

        print('==== 7_M_Ratio_R:  [Mean]   [Stdev]   [ CV% ] ====\n'
              '                    {}      {}     {}'.format(mean_7_M_R,stdev_7_M_R,cv_7_M_R))
        print('==== 8_4_Ratio_R:  [Mean]   [Stdev]   [ CV% ]   ====\n'
              '                    {}      {}     {}'.format(mean_8_4_R,stdev_8_4_R,cv_8_4_R))

        list_for_pass_to_plot =[zone5_C_Ratio_list_B,
                                zone6_R_Ratio_list_B,
                                zone5_C_Ratio_list_G,
                                zone6_R_Ratio_list_G,
                                zone7_M_Ratio_list_R,
                                zone8_4_Ratio_list_R ]
        list_mean_for_pass_to_plot = [mean_5_C_B,
                                      mean_6_R_B,
                                      mean_5_C_G,
                                      mean_6_R_G,
                                      mean_7_M_R,
                                      mean_8_4_R ]

        self.FiveColorCard_SingleLight_plot( list_for_pass_to_plot, list_mean_for_pass_to_plot)

    # 單色光畫圖
    def FiveColorCard_SingleLight_plot(self,list_for_pass_to_plot,list_mean_for_pass_to_plot):

        plt.figure(figsize=(12, 6))
        plt.figure(1)

        # 5_C_B:
        plt.subplot(241)
        data =list_for_pass_to_plot[0]
        bins = np.arange(0.0, 0.6, 0.01)  # fixed bin size
        plt.xlim([min(data) - 0.1, max(data) + 0.1])
        plt.hist(data, bins=bins, color='b', alpha=0.7)
        plt.xlabel('RATIO')
        plt.title('5_C_B RATIO')
        plt.grid(True)

        # 6_R_B:
        plt.subplot(242)
        data =list_for_pass_to_plot[1]
        bins = np.arange(0.0, 0.6, 0.01)  # fixed bin size
        plt.xlim([min(data) - 0.1, max(data) + 0.1])
        plt.hist(data, bins=bins, color='b', alpha=0.7)
        plt.xlabel('RATIO')
        plt.title('6_R_B RATIO')
        plt.grid(True)

        # 5_C_G:
        plt.subplot(243)
        data =list_for_pass_to_plot[2]
        bins = np.arange(0.0, 0.6, 0.01)  # fixed bin size
        plt.xlim([min(data) - 0.1, max(data) + 0.1])
        plt.hist(data, bins=bins, color='G', alpha=0.7)
        plt.xlabel('RATIO')
        plt.title('5_C_G RATIO')
        plt.grid(True)

        # 6_R_G:
        plt.subplot(244)
        data =list_for_pass_to_plot[3]
        bins = np.arange(0.0, 0.6, 0.01)  # fixed bin size
        plt.xlim([min(data), max(data) + 0.05])
        plt.hist(data, bins=bins, color='G', alpha=0.7)
        plt.xlabel('RATIO')
        plt.title('6_R_G RATIO')
        plt.grid(True)

        # 7_M_R:
        plt.subplot(245)
        data =list_for_pass_to_plot[4]
        bins = np.arange(0.0, 0.8, 0.01)  # fixed bin size
        plt.xlim([min(data) - 0.1, max(data) + 0.1])
        plt.hist(data, bins=bins, color='r', alpha=0.7)
        plt.xlabel('RATIO')
        plt.title('7_M_R RATIO')
        plt.grid(True)

        # 8_4_R:
        plt.subplot(246)
        data =list_for_pass_to_plot[5]
        bins = np.arange(0.0, 0.6, 0.01)  # fixed bin size
        plt.xlim([min(data) - 0.1, max(data) + 0.1])
        plt.hist(data, bins=bins, color='r', alpha=0.7)
        plt.xlabel('RATIO')
        plt.title('8_4_R RATIO')
        plt.grid(True)

        plt.subplots_adjust(hspace=0.7)


    # scatter:

        plt.figure(figsize=(12, 6))
        plt.figure(2)

        # 5_C_B:
        plt.subplot(241)

        x = [i for i in range(len(list_for_pass_to_plot[0]))]
        y = list_for_pass_to_plot[0]
        plt.scatter(x, y, c='b', alpha=0.5)
        mean_y = [list_mean_for_pass_to_plot[0]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)
        plt.title('5_C_B')
        plt.ylabel('Ratio')
        plt.grid(True)
        plt.ylim([min(y) -0.05, max(y) + 0.05])

        # 6_R_B:
        plt.subplot(242)

        x = [i for i in range(len(list_for_pass_to_plot[1]))]
        y = list_for_pass_to_plot[1]
        plt.scatter(x, y, c='b', alpha=0.5)
        mean_y = [list_mean_for_pass_to_plot[1]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)
        plt.title('6_R_B')
        plt.ylabel('Ratio')
        plt.grid(True)
        plt.ylim([min(y) - 0.05, max(y) + 0.05])


        # 5_C_G:
        plt.subplot(243)

        x = [i for i in range(len(list_for_pass_to_plot[2]))]
        y = list_for_pass_to_plot[2]
        plt.scatter(x, y, c='G', alpha=0.5)
        mean_y = [list_mean_for_pass_to_plot[2]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)
        plt.title('5_C_G')
        plt.ylabel('Ratio')
        plt.grid(True)
        plt.ylim([min(y) - 0.05, max(y) + 0.05])

        # 6_R_G:
        plt.subplot(244)

        x = [i for i in range(len(list_for_pass_to_plot[3]))]
        y = list_for_pass_to_plot[3]
        plt.scatter(x, y, c='g', alpha=0.5)
        mean_y = [list_mean_for_pass_to_plot[3]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)
        plt.title('G_R_G')
        plt.ylabel('Ratio')
        plt.grid(True)
        plt.ylim([min(y) - 0.05, max(y) + 0.05])

        # 7_M_R:
        plt.subplot(245)

        x = [i for i in range(len(list_for_pass_to_plot[4]))]
        y = list_for_pass_to_plot[4]
        plt.scatter(x, y, c='r', alpha=0.5)
        mean_y = [list_mean_for_pass_to_plot[4]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)
        plt.title('7_M_R')
        plt.ylabel('Ratio')
        plt.grid(True)
        plt.ylim([min(y) - 0.05, max(y) + 0.05])

        # 8_4_R:
        plt.subplot(246)

        x = [i for i in range(len(list_for_pass_to_plot[5]))]
        y = list_for_pass_to_plot[5]
        plt.scatter(x, y, c='r', alpha=0.5)
        mean_y = [list_mean_for_pass_to_plot[5]] * len(x)
        plt.plot(x, mean_y, c='y', alpha=0.5)
        plt.title('8_4_R')
        plt.ylabel('Ratio')
        plt.grid(True)
        plt.ylim([min(y) - 0.05, max(y) + 0.05])

        plt.subplots_adjust( wspace = 0.4,hspace=0.7)



        plt.show()


    def DailyTestAnalysis(self):

        folder_name = 'DailyTest_ProductionLine'

        file_name = 'DataRecord.txt'


        B29 = []
        B30 = []
        G31 = []
        G32 = []
        R33 = []
        R34 = []


        B2B_29 = []
        B2B_30 = []
        G2G_31 = []
        G2G_32 = []
        R2R_33 = []
        R2R_34 = []

        b2b_temp = []
        bb1 = []
        bb2 = []
        bb3 = []
        B_ROI5_Calibration_no1 = []    #B2B -1
        B_ROI6_Reaction_no1    = []    #B2B -2
        B_ROI5_Calibration_no2 = []
        B_ROI6_Reaction_no2    = []
        B_ROI5_Calibration_no3 = []
        B_ROI6_Reaction_no3    = []

        g2g_temp = []
        gg1 = []
        gg2 = []
        gg3 = []
        G_ROI5_Calibration_no1 = []
        G_ROI6_Reaction_no1    = []
        G_ROI5_Calibration_no2 = []
        G_ROI6_Reaction_no2    = []
        G_ROI5_Calibration_no3 = []
        G_ROI6_Reaction_no3    = []

        r2r_temp = []
        rr1 = []
        rr2 = []
        rr3 = []
        R_ROI7_Manual_no1 = []
        R_ROI8_ROI4_no1   = []
        R_ROI7_Manual_no2 = []
        R_ROI8_ROI4_no2   = []
        R_ROI7_Manual_no3 = []
        R_ROI8_ROI4_no3   = []



        ori_path = os.getcwd()
        analysis_folder = ori_path + '/'+ folder_name
        folder_path = os.listdir(analysis_folder)  # 日期資料夾位置
        #print(folder_path)

        for folder in folder_path:
            if folder!='.DS_Store':
                os.chdir(analysis_folder + '/' + folder)
                print('Now in folder: {}'.format(folder))

                folder_path2 = os.listdir(analysis_folder + '/' + folder) # 三隻手機資料夾位置
                for phone_folder in folder_path2:
                    if phone_folder != '.DS_Store':
                        print("第{}手機".format(phone_folder))
                        os.chdir(analysis_folder + '/' + folder+'/'+ phone_folder)
                        if os.stat(file_name).st_size == 0:
                            print('====== 空空如也 ======')

                        else:

                            if phone_folder == "#1":
                                count= 0
                                with open(file_name, 'r') as f:
                                    for i in f.readlines():
                                        count += 1
                                        if '$Production_Line_Calibration_FiveColor_Blue_LightB2B' in i:
                                            # 當找到要的資料時 做字串分類，且分兩個
                                            b2b_temp.append(linecache.getline(analysis_folder + '/' + folder + '/'+ phone_folder + '/' + file_name,count +1))
                                            for ii in b2b_temp:
                                                b= ii.splitlines() #一個LIST裡面放一個字串
                                                for iii in b:
                                                    bb1= iii.split(' ')
                                                    print(bb1)
                                                B_ROI5_Calibration_no1.append(bb1[0])
                                                B_ROI6_Reaction_no1.append((bb1[1]))

                                        elif '$Production_Line_Calibration_FiveColor_Green_LightG2G' in i:
                                            g2g_temp.append(linecache.getline(analysis_folder + '/' + folder + '/' + phone_folder + '/' + file_name,count + 1))
                                            for ii in g2g_temp:
                                                g = ii.splitlines()  # 一個LIST裡面放一個字串
                                                for iii in g:
                                                    gg1 = iii.split(' ')
                                                    print(gg1)
                                                G_ROI5_Calibration_no1.append(gg1[0])
                                                G_ROI6_Reaction_no1.append((gg1[1]))

                                        elif '$Production_Line_Calibration_FiveColor_Red_LightR2R' in i:
                                            r2r_temp.append(linecache.getline(analysis_folder + '/' + folder + '/' + phone_folder + '/' + file_name,count + 1))
                                            for ii in r2r_temp:
                                                r = ii.splitlines()  # 一個LIST裡面放一個字串
                                                for iii in r:
                                                    rr1 = iii.split(' ')
                                                    print(rr1)
                                                R_ROI7_Manual_no1.append(rr1[0])
                                                R_ROI8_ROI4_no1.append((rr1[1]))


                            elif phone_folder == "#2":
                                count = 0
                                with open(file_name, 'r') as f:
                                    for i in f.readlines():
                                        count += 1
                                        if '$Production_Line_Calibration_FiveColor_Blue_LightB2B' in i:
                                            # 當找到要的資料時 做字串分類，且分兩個
                                            b2b_temp.append(linecache.getline(analysis_folder + '/' + folder + '/'+ phone_folder + '/' + file_name,count +1))
                                            for ii in b2b_temp:
                                                b= ii.splitlines()
                                                for iii in b:
                                                    bb2= iii.split(' ')
                                                    print(bb2)
                                                    B_ROI5_Calibration_no2.append(bb2[0])
                                                    B_ROI6_Reaction_no2.append(bb2[1])

                                        elif '$Production_Line_Calibration_FiveColor_Green_LightG2G' in i:
                                            g2g_temp.append(linecache.getline(analysis_folder + '/' + folder + '/' + phone_folder + '/' + file_name,count + 1))
                                            for ii in g2g_temp:
                                                g = ii.splitlines()  # 一個LIST裡面放一個字串
                                                for iii in g:
                                                    gg2 = iii.split(' ')
                                                    print(gg2)
                                                G_ROI5_Calibration_no2.append(gg2[0])
                                                G_ROI6_Reaction_no2.append((gg2[1]))

                                        elif '$Production_Line_Calibration_FiveColor_Red_LightR2R' in i:
                                            r2r_temp.append(linecache.getline(analysis_folder + '/' + folder + '/' + phone_folder + '/' + file_name,count + 1))
                                            for ii in r2r_temp:
                                                r = ii.splitlines()  # 一個LIST裡面放一個字串
                                                for iii in r:
                                                    rr2 = iii.split(' ')
                                                    print(rr2)
                                                R_ROI7_Manual_no2.append(rr2[0])
                                                R_ROI8_ROI4_no2.append((rr2[1]))


                            elif phone_folder == "#3":
                                count = 0
                                with open(file_name, 'r') as f:
                                    for i in f.readlines():
                                        count += 1
                                        if '$Production_Line_Calibration_FiveColor_Blue_LightB2B' in i:
                                            # 當找到要的資料時 做字串分類，且分兩個
                                            b2b_temp.append(linecache.getline(analysis_folder + '/' + folder + '/'+ phone_folder + '/' + file_name,count +1))
                                            for ii in b2b_temp:
                                                b= ii.splitlines()
                                                for iii in b:
                                                    bb3= iii.split(' ')
                                                    print(bb3)
                                                    B_ROI5_Calibration_no3.append(bb3[0])
                                                    B_ROI6_Reaction_no3.append(bb3[1])
                                        elif '$Production_Line_Calibration_FiveColor_Green_LightG2G' in i:
                                            g2g_temp.append(linecache.getline(analysis_folder + '/' + folder + '/' + phone_folder + '/' + file_name,count + 1))
                                            for ii in g2g_temp:
                                                g = ii.splitlines()  # 一個LIST裡面放一個字串
                                                for iii in g:
                                                    gg3 = iii.split(' ')
                                                    print(gg3)
                                                G_ROI5_Calibration_no3.append(gg3[0])
                                                G_ROI6_Reaction_no3.append((gg3[1]))

                                        elif '$Production_Line_Calibration_FiveColor_Red_LightR2R' in i:
                                            r2r_temp.append(linecache.getline(analysis_folder + '/' + folder + '/' + phone_folder + '/' + file_name,count + 1))
                                            for ii in r2r_temp:
                                                r = ii.splitlines()  # 一個LIST裡面放一個字串
                                                for iii in r:
                                                    rr3 = iii.split(' ')
                                                    print(rr3)
                                                R_ROI7_Manual_no3.append(rr3[0])
                                                R_ROI8_ROI4_no3.append((rr3[1]))

                    b2b_temp = []
                    g2g_temp = []
                    r2r_temp = []
                print('')
            bb1 = []
            bb2 = []
            bb3 = []
            gg1 = []
            gg2 = []
            gg3 = []
            rr1 = []
            rr2 = []
            rr3 = []
        os.chdir(ori_path)

        print("=========  B  =============")
        print(B_ROI5_Calibration_no1)
        print(B_ROI6_Reaction_no1)
        print(B_ROI5_Calibration_no2)
        print(B_ROI6_Reaction_no2)
        print(B_ROI5_Calibration_no3)
        print(B_ROI6_Reaction_no3)
        print("=========  G  =============")
        print(G_ROI5_Calibration_no1)
        print(G_ROI6_Reaction_no1)
        print(G_ROI5_Calibration_no2)
        print(G_ROI6_Reaction_no2)
        print(G_ROI5_Calibration_no3)
        print(G_ROI6_Reaction_no3)
        print("=========  R  =============")
        print(R_ROI7_Manual_no1)
        print(R_ROI8_ROI4_no1)
        print(R_ROI7_Manual_no2)
        print(R_ROI8_ROI4_no2)
        print(R_ROI7_Manual_no3)
        print(R_ROI8_ROI4_no3)



        # 以下是要轉pandas準備：
        for i in range(len(B_ROI5_Calibration_no1)):
            B29.append(B_ROI5_Calibration_no1[i])
            B29.append(B_ROI5_Calibration_no2[i])
            B29.append(B_ROI5_Calibration_no3[i])
            B2B_29.append(B29)
            B29 = []
        print(B2B_29)

        for i in range(len(B_ROI6_Reaction_no1)):
            B30.append(B_ROI6_Reaction_no1[i])
            B30.append(B_ROI6_Reaction_no2[i])
            B30.append(B_ROI6_Reaction_no3[i])
            B2B_30.append(B30)
            B30 = []
        print(B2B_30)

        for i in range(len(G_ROI5_Calibration_no1)):
            G31.append(G_ROI5_Calibration_no1[i])
            G31.append(G_ROI5_Calibration_no2[i])
            G31.append(G_ROI5_Calibration_no3[i])
            G2G_31.append(G31)
            G31 = []
        print(G2G_31)

        for i in range(len(G_ROI6_Reaction_no1)):
            G32.append(G_ROI6_Reaction_no1[i])
            G32.append(G_ROI6_Reaction_no2[i])
            G32.append(G_ROI6_Reaction_no3[i])
            G2G_32.append(G32)
            G32 = []
        print(G2G_32)

        for i in range(len(R_ROI7_Manual_no1)):
            R33.append(R_ROI7_Manual_no1[i])
            R33.append(R_ROI7_Manual_no2[i])
            R33.append(R_ROI7_Manual_no3[i])
            R2R_33.append(R33)
            R33 = []
        print(R2R_33)

        for i in range(len(R_ROI8_ROI4_no1)):
            R34.append(R_ROI8_ROI4_no1[i])
            R34.append(R_ROI8_ROI4_no2[i])
            R34.append(R_ROI8_ROI4_no3[i])
            R2R_34.append(R34)
            R34 = []
        print(R2R_34)

        array29 = np.asarray(B2B_29)
        array30 = np.asarray(B2B_30)
        array31 = np.asarray(G2G_31)
        array32 = np.asarray(G2G_32)
        array33 = np.asarray(R2R_33)
        array34 = np.asarray(R2R_34)


        df29 = pd.DataFrame(array29, columns= ['#1', '#2', '#3'])
        df30 = pd.DataFrame(array30, columns= ['#1', '#2', '#3'])
        df31 = pd.DataFrame(array31, columns= ['#1', '#2', '#3'])
        df32 = pd.DataFrame(array32, columns= ['#1', '#2', '#3'])
        df33 = pd.DataFrame(array33, columns= ['#1', '#2', '#3'])
        df34 = pd.DataFrame(array34, columns= ['#1', '#2', '#3'])




        # 確認有無檔案：
        status= True
        while status == True:
            keyin_filename = input('輸入 年月日 作為excel file 或 0 則是原本excel添加資料 ： ')
            if keyin_filename == '0':

                    wb = load_workbook('WaHaHa.xlsx')  # 先讀取已存在的檔案，並獲取row數量
                    sheet = wb.worksheets[0]
                    row_count = sheet.max_row
                    writer = pd.ExcelWriter('WaHaHa.xlsx') # 建立excelwriter 做寫入動作
                    writer.book = wb                                               # 把前面讀取的檔案傳給 wirter （不傳writer會當作是新檔案直接複寫)
                    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)  # 連sheet也要傳給 wirter 不傳writer會當作是新檔案直接複寫)
                    df29.to_excel(writer, '29', header= None, startrow= row_count)
                    df30.to_excel(writer, '30', header= None, startrow= row_count)
                    df31.to_excel(writer, '31', header= None, startrow= row_count)
                    df32.to_excel(writer, '32', header= None, startrow= row_count)
                    df33.to_excel(writer, '33', header= None, startrow= row_count)
                    df34.to_excel(writer, '34', header= None, startrow= row_count)
                    writer.save()
                    status = False
            else:
                writer = pd.ExcelWriter('WaHaHa_{}.xlsx'.format(keyin_filename))
                df29.to_excel(writer, '29')
                df30.to_excel(writer, '30')
                df31.to_excel(writer, '31')
                df32.to_excel(writer, '32')
                df33.to_excel(writer, '33')
                df34.to_excel(writer, '34')
                writer.save()
                status= False










if __name__ =='__main__':
    ix = ixensor()
    #ix.Uniformity  # 算均勻度
    # ix.Contour_3d() # 可視化
    #ix.DarkSignal() #算暗訊號
    # ix.Signal_Click()
    #ix.Folder_Sort()  # 資料夾整理
    # ix.over_exposure_comapre()
    # ix.Signal_curve_A1C_lIPID()   # potc訊號曲線


    # ix.FiveColorCard_Signal_Comfirm() # 五色卡 白光
    # ix.FiveColorCard_SingleLight_statictics() #五色卡單色光
    ix.DailyTestAnalysis()


    print(threading.current_thread())


